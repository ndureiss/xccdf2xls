#! /usr/bin/python3
import argparse
from os import path
from glob import glob
from xml.etree.ElementTree import parse, ParseError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

####################################################################################################################
#  Python script wrote by ndureiss
#  Please write a Github issue in case of bugs or evolution request
#
#  Wrote for python3
#  Dependencies to :
#    - openpyxl

####################################################################################################################


def parsingFile(filePath):
    # Parsing a file as an XML
    # filepath: the path of file to parse
    print("Reading file {} ..".format(filePath))
    try:
        tree = parse(filePath)
    except ParseError as err:
        print("\033[91mError parsing {}\033[0m".format(filePath))
        quit()
    return tree


def getXMLNS(rootElem):
    return rootElem.tag[1:].partition("}")[0] if rootElem.tag[0] == "{" else None


def computeRefResult(ruleResults):
    if "fail" in ruleResults:
        return "fail"
    if "unknown" in ruleResults:
        return "unknown"
    if "error" in ruleResults:
        return "error"
    if "pass" in ruleResults:
        return "pass"
    if "unchecked" in ruleResults:
        return "unchecked"
    return "notselected"


def getBgColor(value):
    if "fail" in value:
        return "DA9694"
    if "unknown" in value:
        return "FABF8F"
    if "pass" in value:
        return "C4C79B"
    if "unchecked" in value:
        return "BFBFBF"
    return "FFFFFF"


def getFontColor(value):
    if "fail" in value:
        return "632523"
    if "unknown" in value:
        return "974706"
    if "pass" in value:
        return "4F6228"
    if "unchecked" in value:
        return "808080"
    if "notselected" in value:
        return "000000"
    return "632523"


def getResFontColor(value):
    if value > 0.95:
        return "009900"
    return "CC0000"


def addKeyValuePairToDict(key, value, dictionary):
    # Add a (Key, Value) pair to dictionary
    # key: the key of the pair
    # value: the value of the pair
    # dictionary: the destination dictionary
    if key in dictionary:
        dictionary[key].update(value)
    else:
        dictionary[key] = value


def flatDictKeys(dictionary):
    # Flat one level dictionary by keys
    # dictionary: the one to flat
    res = list()
    for ref, rules in dictionary.items():
        res.append("[REF]: {}".format(ref))
        [res.append(r) for r in rules.keys()]
    return res


def flatDictValues(dictionary):
    # Flat one level dictionary by values
    # dictionary: the one to flat
    res = list()
    for ref, rules in dictionary.items():
        results = rules.values()
        res.append(computeRefResult(list(results)))
        [res.append(r) for r in results]
    return res


def formatCell(cell, bold=False, align="left", fullBorders=False, bgColor="FFFFFF", color="000000", nbFormat=''):
    # Applying a style to a cell
    # cell: the targeted cell
    # bold: bold text
    # align: text alignement in cell
    # fullBorders: applying borders to top and bottom
    # bgColor: background color to apply, hexadecimal format
    # color: font color to apply, hexadecimal format
    # nbFormat: number format to apply if necessary
    _s = Side(style='thin')
    cell.font = Font(bold=bold, color=color)
    cell.fill = PatternFill(fgColor=bgColor, fill_type="solid")
    cell.border = Border(left=_s, right=_s, top=_s,
                         bottom=_s) if fullBorders else Border(left=_s, right=_s)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.number_format = nbFormat


def autosizeWorksheet(worksheet):
    # Autosize columns of the worksheet
    # TODO: Not optimal, try to improve this
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            colLetter = cell.column_letter
            dims[colLetter] = max(
                (dims.get(colLetter, 0), len(str(cell.value))))
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value


def computeTestSuccessRate(worksheet, row, lastCol):
    # Compute success rate of a given test over machines
    # worksheet: the test result worksheet
    # row: the test row index
    # lastCol: the last machine column index
    return sum([1 if worksheet.cell(row=row, column=c).value
                in {"pass", "notselected", "unchecked"} else 0
                for c in range(2, lastCol)])/(lastCol-2)


def xccdf2json(filePath, grouped=False, group="UNREFERENCED"):
    # Convert XCCDF result xml(s) to a JSON object
    # filePath: absolute path to find XML files
    # grouped: boolean to indicate if rule results must be grouped
    # groupName: group name for grouping func
    mainDict = dict()
    for machineId, file in enumerate(glob(filePath)):
        root = parsingFile(file).getroot()
        xmlns = getXMLNS(root)
        # Map Rule IDs to Rule results
        ruleDict = dict()
        for rr in root.iter("{%s}rule-result" % xmlns):
            result = rr.find("{%s}result" % xmlns)
            ruleDict[rr.get("idref")] = result.text

        refDict = ruleDict
        # Grouping means mapping Rule refs to Rule IDs with Rule results
        if (grouped):
            refDict = dict()
            for elem in root.iter("{%s}Rule" % xmlns):
                tmp = dict()
                tmp[elem.get("id")] = ruleDict.get(elem.get("id"))
                for r in elem.iter("{%s}reference" % xmlns):
                    key = r.text if r.get("href") == group else "UNREFERENCED"
                    addKeyValuePairToDict(key, tmp, refDict)
                if len(list(elem.iter("{%s}reference" % xmlns))) == 0:
                    addKeyValuePairToDict("UNREFERENCED", tmp, refDict)
            refDict = dict(sorted(refDict.items(), key=lambda x: x[0].lower()))

        testResult = root.find("{%s}TestResult" % xmlns)
        mainDict[testResult.find("{%s}target" % xmlns).text] = {
            "test_results": refDict,
            "score": float(testResult.find("{%s}score" % xmlns).text) / float(testResult.find("{%s}score" % xmlns).get("maximum"))
        }
    return mainDict


####################################################################################################################
grouped = None
groupName = None
parser = argparse.ArgumentParser()
parser.add_argument(
    "-p", "--path", help="XML files, alias (*) accepted, must be quoted, default \"*\"", type=str, default="*")
parser.add_argument(
    "-g", "--group", help="reference to group result by, default null", type=str, default="")
parser.add_argument(
    "-o", "--output", help="output file name, default result.xlsx", type=str, default="result.xlsx")
args = parser.parse_args()
if args.group:
    grouped = True
    groupName = args.group

files = "{}/{}.xml".format(path.abspath(path.dirname(parser.prog)), args.path)
if len(glob(files)) == 0:
    print("\033[91mNone file found!\033[0m")
    quit()

res = xccdf2json(files, grouped, groupName)

# Create Workbook and initialize Results WorkSheet
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Results"

lastMachineCol = 2
for machineNum, (machineName, mapping) in enumerate(res.items()):
    if machineNum == 0:
        # Fill first column
        firstCol = flatDictKeys(mapping["test_results"]) if grouped else list(
            mapping["test_results"].keys())
        lastRow = 2
        for rowIndex, rowValue in enumerate(firstCol):
            cell = worksheet.cell(row=rowIndex+2, column=1)
            cell.value = rowValue
            formatCell(cell)
            lastRow += 1
            if "[REF]" in rowValue:
                formatCell(cell, bold=True, bgColor="DCE6F1", fullBorders=True)
            else:
                worksheet.row_dimensions[rowIndex+2].hidden = True
                worksheet.row_dimensions[rowIndex+2].outlineLevel = 1
        cell = worksheet.cell(row=lastRow, column=1)
        cell.value = "SCAP SCORE"
        formatCell(cell, bold=True, align="right", fullBorders=True)

    # Fill machine column
    cell = worksheet.cell(row=1, column=machineNum+2)
    cell.value = machineName
    formatCell(cell, bold=True, align="center", fullBorders=True)
    machineCol = flatDictValues(mapping["test_results"]) if grouped else list(
        mapping["test_results"].values())
    for rowIndex, rowValue in enumerate(machineCol):
        cell = worksheet.cell(row=rowIndex+2, column=machineNum+2)
        cell.value = rowValue
        formatCell(cell)
        if "[REF]" in worksheet.cell(row=rowIndex+2, column=1).value:
            formatCell(cell, bold=True, bgColor=getBgColor(cell.value),
                       fullBorders=True, color=getFontColor(cell.value))

    # Fill machine SCAP computed score
    cell = worksheet.cell(row=lastRow, column=machineNum+2)
    cell.value = mapping["score"]
    formatCell(cell, bold=True, fullBorders=True,
               align="right", nbFormat='0.00%')
    lastMachineCol += 1

# Compute Test Achievement
for r in range(2, lastRow):
    cell = worksheet.cell(row=r, column=lastMachineCol)
    cell.value = computeTestSuccessRate(worksheet, r, lastMachineCol)
    if "[REF]" in worksheet.cell(row=r, column=1).value:
        formatCell(cell, bold=True, align="right", nbFormat='0%',
                   color=getResFontColor(cell.value))
    else:
        formatCell(cell, align="right", fullBorders=True, nbFormat='0%')


autosizeWorksheet(worksheet)
worksheet.freeze_panes = worksheet["B2"]

# Save file and quit
workbook.save(args.output)
print("\033[92mSuccessfully merged and convert!\033[0m")
